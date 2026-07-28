"""PDF and Excel exports for the NSSA statutory report."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import (
    ParagraphStyle,
    getSampleStyleSheet,
)
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


ZERO = Decimal("0.00")


class NssaExportService:
    """Generate NSSA statutory PDF and Excel reports."""

    COMPANY_NAME = "BUYOH (Pvt) Ltd"
    SYSTEM_NAME = "BUYOH Payroll System"
    COMPANY_LOCATION = "Harare, Zimbabwe"

    @staticmethod
    def _money(value):
        """Return a formatted monetary value."""

        if value is None:
            value = ZERO

        return f"${Decimal(str(value)):,.2f}"

    @staticmethod
    def _employee_name(record):
        """Return an employee's full name."""

        employee = record.employee

        return (
            f"{employee.first_name} "
            f"{employee.last_name}"
        ).strip()

    @staticmethod
    def _generated_by_name(user):
        """Return the best available user display name."""

        if user is None:
            return "Unknown User"

        first_name = getattr(
            user,
            "first_name",
            None,
        )

        last_name = getattr(
            user,
            "last_name",
            None,
        )

        full_name = " ".join(
            part
            for part in [
                first_name,
                last_name,
            ]
            if part
        ).strip()

        if full_name:
            return full_name

        return getattr(
            user,
            "username",
            "Unknown User",
        )

    @classmethod
    def generate_pdf(
        cls,
        selected_period,
        rows,
        totals,
        generated_by,
        search_term="",
        selected_department=None,
    ):
        """Generate a landscape NSSA PDF report."""

        buffer = BytesIO()
        generated_at = datetime.now()

        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title=(
                "NSSA Report - "
                f"{selected_period.period_name}"
            ),
            author=cls.SYSTEM_NAME,
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "NssaTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=22,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#111827"),
            spaceAfter=3,
        )

        company_style = ParagraphStyle(
            "NssaCompany",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#1F2937"),
        )

        metadata_style = ParagraphStyle(
            "NssaMetadata",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#4B5563"),
        )

        right_metadata_style = ParagraphStyle(
            "NssaRightMetadata",
            parent=metadata_style,
            alignment=TA_RIGHT,
        )

        summary_label_style = ParagraphStyle(
            "NssaSummaryLabel",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#6B7280"),
        )

        summary_value_style = ParagraphStyle(
            "NssaSummaryValue",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#111827"),
        )

        elements = []

        left_header = [
            Paragraph(
                cls.SYSTEM_NAME,
                company_style,
            ),
            Paragraph(
                "NSSA Statutory Report",
                title_style,
            ),
            Paragraph(
                (
                    f"{cls.COMPANY_NAME}<br/>"
                    f"{cls.COMPANY_LOCATION}"
                ),
                metadata_style,
            ),
        ]

        right_header_lines = [
            (
                "<b>Payroll Period:</b> "
                f"{selected_period.period_name}"
            ),
            (
                "<b>Payment Date:</b> "
                f"{selected_period.payment_date.strftime('%d %B %Y')}"
            ),
            (
                "<b>Generated:</b> "
                f"{generated_at.strftime('%d %B %Y %H:%M')}"
            ),
            (
                "<b>Generated By:</b> "
                f"{cls._generated_by_name(generated_by)}"
            ),
        ]

        if selected_department is not None:
            right_header_lines.append(
                (
                    "<b>Department:</b> "
                    f"{selected_department.name}"
                )
            )

        if search_term:
            right_header_lines.append(
                (
                    "<b>Search Filter:</b> "
                    f"{search_term}"
                )
            )

        right_header = Paragraph(
            "<br/>".join(right_header_lines),
            right_metadata_style,
        )

        header_table = Table(
            [[left_header, right_header]],
            colWidths=[
                170 * mm,
                95 * mm,
            ],
        )

        header_table.setStyle(
            TableStyle(
                [
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP",
                    ),
                    (
                        "ALIGN",
                        (1, 0),
                        (1, 0),
                        "RIGHT",
                    ),
                ]
            )
        )

        elements.append(header_table)
        elements.append(Spacer(1, 5 * mm))

        summary_data = [
            [
                Paragraph(
                    "Employees",
                    summary_label_style,
                ),
                Paragraph(
                    "Gross Payroll",
                    summary_label_style,
                ),
                Paragraph(
                    "Employee NSSA",
                    summary_label_style,
                ),
                Paragraph(
                    "Employer NSSA",
                    summary_label_style,
                ),
                Paragraph(
                    "Total NSSA",
                    summary_label_style,
                ),
            ],
            [
                Paragraph(
                    str(totals["employee_count"]),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["gross_pay"]
                    ),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["employee_nssa"]
                    ),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["employer_nssa"]
                    ),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["total_nssa"]
                    ),
                    summary_value_style,
                ),
            ],
        ]

        summary_table = Table(
            summary_data,
            colWidths=[
                53 * mm,
                53 * mm,
                53 * mm,
                53 * mm,
                53 * mm,
            ],
        )

        summary_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        colors.HexColor("#F3F4F6"),
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.6,
                        colors.HexColor("#D1D5DB"),
                    ),
                    (
                        "INNERGRID",
                        (0, 0),
                        (-1, -1),
                        0.3,
                        colors.HexColor("#E5E7EB"),
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                ]
            )
        )

        elements.append(summary_table)
        elements.append(Spacer(1, 5 * mm))

        table_data = [
            [
                "Employee Number",
                "Employee",
                "Department",
                "Gross Pay",
                "Employee NSSA",
                "Employer NSSA",
                "Total Contribution",
            ]
        ]

        for row in rows:
            record = row["record"]

            table_data.append(
                [
                    record.employee.employee_number,
                    cls._employee_name(record),
                    record.employee.department.name,
                    cls._money(
                        row["gross_pay"]
                    ),
                    cls._money(
                        row["employee_nssa"]
                    ),
                    cls._money(
                        row["employer_nssa"]
                    ),
                    cls._money(
                        row["total_nssa"]
                    ),
                ]
            )

        table_data.append(
            [
                "REPORT TOTALS",
                "",
                "",
                cls._money(
                    totals["gross_pay"]
                ),
                cls._money(
                    totals["employee_nssa"]
                ),
                cls._money(
                    totals["employer_nssa"]
                ),
                cls._money(
                    totals["total_nssa"]
                ),
            ]
        )

        nssa_table = Table(
            table_data,
            repeatRows=1,
            colWidths=[
                30 * mm,
                48 * mm,
                45 * mm,
                34 * mm,
                34 * mm,
                34 * mm,
                38 * mm,
            ],
        )

        last_row = len(table_data) - 1

        nssa_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1F2937"),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white,
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, 0),
                        7.5,
                    ),
                    (
                        "ALIGN",
                        (3, 0),
                        (-1, -1),
                        "RIGHT",
                    ),
                    (
                        "ALIGN",
                        (0, 0),
                        (2, -1),
                        "LEFT",
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.35,
                        colors.HexColor("#D1D5DB"),
                    ),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, last_row - 1),
                        [
                            colors.white,
                            colors.HexColor("#F9FAFB"),
                        ],
                    ),
                    (
                        "BACKGROUND",
                        (0, last_row),
                        (-1, last_row),
                        colors.HexColor("#E5E7EB"),
                    ),
                    (
                        "FONTNAME",
                        (0, last_row),
                        (-1, last_row),
                        "Helvetica-Bold",
                    ),
                    (
                        "SPAN",
                        (0, last_row),
                        (2, last_row),
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                ]
            )
        )

        elements.append(nssa_table)
        elements.append(Spacer(1, 5 * mm))

        average_data = [
            [
                "Average Employee NSSA",
                cls._money(
                    totals[
                        "average_employee_nssa"
                    ]
                ),
                "Average Employer NSSA",
                cls._money(
                    totals[
                        "average_employer_nssa"
                    ]
                ),
                "Average Total Contribution",
                cls._money(
                    totals[
                        "average_total_nssa"
                    ]
                ),
            ]
        ]

        average_table = Table(
            average_data,
            colWidths=[
                40 * mm,
                30 * mm,
                40 * mm,
                30 * mm,
                44 * mm,
                30 * mm,
            ],
            hAlign="LEFT",
        )

        average_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        colors.HexColor("#F9FAFB"),
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.35,
                        colors.HexColor("#D1D5DB"),
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "Helvetica",
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (0, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTNAME",
                        (2, 0),
                        (2, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTNAME",
                        (4, 0),
                        (4, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "ALIGN",
                        (1, 0),
                        (1, 0),
                        "RIGHT",
                    ),
                    (
                        "ALIGN",
                        (3, 0),
                        (3, 0),
                        "RIGHT",
                    ),
                    (
                        "ALIGN",
                        (5, 0),
                        (5, 0),
                        "RIGHT",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, -1),
                        7,
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                ]
            )
        )

        elements.append(average_table)

        def draw_footer(canvas, doc):
            """Draw a footer and page number."""

            canvas.saveState()

            page_width, _ = landscape(A4)

            canvas.setStrokeColor(
                colors.HexColor("#D1D5DB")
            )
            canvas.setLineWidth(0.4)

            canvas.line(
                12 * mm,
                8 * mm,
                page_width - 12 * mm,
                8 * mm,
            )

            canvas.setFont(
                "Helvetica",
                7,
            )
            canvas.setFillColor(
                colors.HexColor("#6B7280")
            )

            canvas.drawString(
                12 * mm,
                4.5 * mm,
                (
                    f"{cls.SYSTEM_NAME} | "
                    f"{selected_period.period_name}"
                ),
            )

            canvas.drawRightString(
                page_width - 12 * mm,
                4.5 * mm,
                f"Page {doc.page}",
            )

            canvas.restoreState()

        document.build(
            elements,
            onFirstPage=draw_footer,
            onLaterPages=draw_footer,
        )

        buffer.seek(0)

        return buffer

    @classmethod
    def generate_excel(
        cls,
        selected_period,
        rows,
        totals,
        generated_by,
        search_term="",
        selected_department=None,
    ):
        """Generate a formatted NSSA Excel workbook."""

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "NSSA Report"

        generated_at = datetime.now()

        dark_fill = PatternFill(
            fill_type="solid",
            fgColor="1F2937",
        )

        blue_fill = PatternFill(
            fill_type="solid",
            fgColor="EAF2FF",
        )

        light_fill = PatternFill(
            fill_type="solid",
            fgColor="F3F4F6",
        )

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="DCFCE7",
        )

        white_font = Font(
            color="FFFFFF",
            bold=True,
        )

        title_font = Font(
            size=18,
            bold=True,
            color="111827",
        )

        label_font = Font(
            bold=True,
            color="374151",
        )

        value_font = Font(
            bold=True,
            color="111827",
        )

        thin_side = Side(
            style="thin",
            color="D1D5DB",
        )

        table_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        worksheet.merge_cells(
            "A1:G1"
        )
        worksheet["A1"] = cls.SYSTEM_NAME
        worksheet["A1"].font = title_font

        worksheet.merge_cells(
            "A2:G2"
        )
        worksheet["A2"] = "NSSA Statutory Report"
        worksheet["A2"].font = Font(
            size=14,
            bold=True,
            color="111827",
        )

        worksheet.merge_cells(
            "A3:G3"
        )
        worksheet["A3"] = (
            f"{cls.COMPANY_NAME} | "
            f"{cls.COMPANY_LOCATION}"
        )
        worksheet["A3"].font = Font(
            italic=True,
            color="6B7280",
        )

        metadata_rows = [
            (
                "Payroll Period",
                selected_period.period_name,
            ),
            (
                "Payment Date",
                selected_period.payment_date.strftime(
                    "%d %B %Y"
                ),
            ),
            (
                "Generated",
                generated_at.strftime(
                    "%d %B %Y %H:%M"
                ),
            ),
            (
                "Generated By",
                cls._generated_by_name(
                    generated_by
                ),
            ),
        ]

        if selected_department is not None:
            metadata_rows.append(
                (
                    "Department",
                    selected_department.name,
                )
            )

        if search_term:
            metadata_rows.append(
                (
                    "Search Filter",
                    search_term,
                )
            )

        metadata_start_row = 5

        for offset, (label, value) in enumerate(
            metadata_rows
        ):
            row_number = (
                metadata_start_row
                + offset
            )

            worksheet.cell(
                row=row_number,
                column=1,
                value=label,
            ).font = label_font

            worksheet.cell(
                row=row_number,
                column=2,
                value=value,
            )

        summary_start_row = (
            metadata_start_row
            + len(metadata_rows)
            + 2
        )

        summary_headers = [
            "Employees",
            "Gross Payroll",
            "Employee NSSA",
            "Employer NSSA",
            "Total NSSA",
        ]

        summary_values = [
            totals["employee_count"],
            totals["gross_pay"],
            totals["employee_nssa"],
            totals["employer_nssa"],
            totals["total_nssa"],
        ]

        for column, heading in enumerate(
            summary_headers,
            start=1,
        ):
            header_cell = worksheet.cell(
                row=summary_start_row,
                column=column,
                value=heading,
            )

            header_cell.fill = blue_fill
            header_cell.font = label_font
            header_cell.border = table_border
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            value_cell = worksheet.cell(
                row=summary_start_row + 1,
                column=column,
                value=summary_values[
                    column - 1
                ],
            )

            value_cell.fill = light_fill
            value_cell.font = value_font
            value_cell.border = table_border
            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            if column >= 2:
                value_cell.number_format = (
                    '$#,##0.00'
                )

        table_header_row = (
            summary_start_row + 4
        )

        headings = [
            "Employee Number",
            "Employee",
            "Department",
            "Gross Pay",
            "Employee NSSA",
            "Employer NSSA",
            "Total Contribution",
        ]

        for column, heading in enumerate(
            headings,
            start=1,
        ):
            cell = worksheet.cell(
                row=table_header_row,
                column=column,
                value=heading,
            )

            cell.fill = dark_fill
            cell.font = white_font
            cell.border = table_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        data_start_row = table_header_row + 1

        for row_offset, row in enumerate(
            rows
        ):
            row_number = (
                data_start_row
                + row_offset
            )

            record = row["record"]

            values = [
                record.employee.employee_number,
                cls._employee_name(record),
                record.employee.department.name,
                row["gross_pay"],
                row["employee_nssa"],
                row["employer_nssa"],
                row["total_nssa"],
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )

                cell.border = table_border
                cell.alignment = Alignment(
                    vertical="center",
                )

                if column >= 4:
                    cell.number_format = (
                        '$#,##0.00'
                    )
                    cell.alignment = Alignment(
                        horizontal="right",
                        vertical="center",
                    )

        totals_row = (
            data_start_row
            + len(rows)
        )

        totals_values = [
            "REPORT TOTALS",
            "",
            "",
            totals["gross_pay"],
            totals["employee_nssa"],
            totals["employer_nssa"],
            totals["total_nssa"],
        ]

        for column, value in enumerate(
            totals_values,
            start=1,
        ):
            cell = worksheet.cell(
                row=totals_row,
                column=column,
                value=value,
            )

            cell.fill = green_fill
            cell.font = Font(
                bold=True,
                color="111827",
            )
            cell.border = table_border
            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column <= 3
                    else "right"
                ),
                vertical="center",
            )

            if column >= 4:
                cell.number_format = (
                    '$#,##0.00'
                )

        worksheet.merge_cells(
            start_row=totals_row,
            start_column=1,
            end_row=totals_row,
            end_column=3,
        )

        averages_start_row = (
            totals_row + 3
        )

        worksheet.cell(
            row=averages_start_row,
            column=1,
            value="NSSA Averages",
        ).font = Font(
            size=11,
            bold=True,
            color="374151",
        )

        averages = [
            (
                "Average Employee NSSA",
                totals[
                    "average_employee_nssa"
                ],
            ),
            (
                "Average Employer NSSA",
                totals[
                    "average_employer_nssa"
                ],
            ),
            (
                "Average Total Contribution",
                totals[
                    "average_total_nssa"
                ],
            ),
        ]

        for offset, (label, value) in enumerate(
            averages,
            start=1,
        ):
            row_number = (
                averages_start_row
                + offset
            )

            label_cell = worksheet.cell(
                row=row_number,
                column=1,
                value=label,
            )

            value_cell = worksheet.cell(
                row=row_number,
                column=2,
                value=value,
            )

            label_cell.fill = light_fill
            label_cell.font = label_font
            label_cell.border = table_border

            value_cell.border = table_border
            value_cell.number_format = (
                '$#,##0.00'
            )
            value_cell.alignment = Alignment(
                horizontal="right",
            )

        column_widths = {
            1: 20,
            2: 27,
            3: 26,
            4: 17,
            5: 18,
            6: 18,
            7: 20,
        }

        for column, width in (
            column_widths.items()
        ):
            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = width

        worksheet.freeze_panes = (
            f"A{data_start_row}"
        )

        if rows:
            worksheet.auto_filter.ref = (
                f"A{table_header_row}:"
                f"G{totals_row - 1}"
            )

        worksheet.sheet_view.showGridLines = False

        worksheet.page_setup.orientation = (
            "landscape"
        )
        worksheet.page_setup.paperSize = (
            worksheet.PAPERSIZE_A4
        )
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        worksheet.print_title_rows = (
            f"{table_header_row}:"
            f"{table_header_row}"
        )

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return buffer
