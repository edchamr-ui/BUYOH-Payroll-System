"""PDF and Excel export services for payroll reports."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


ZERO = Decimal("0.00")


class ReportExportService:
    """Generate downloadable payroll-summary reports."""

    COMPANY_NAME = "BUYOH (Pvt) Ltd"
    SYSTEM_NAME = "BUYOH Payroll System"
    COMPANY_LOCATION = "Harare, Zimbabwe"

    @staticmethod
    def _money(value):
        """Return a consistently formatted currency value."""

        if value is None:
            value = ZERO

        return f"${Decimal(str(value)):,.2f}"

    @staticmethod
    def _safe_text(value):
        """Return a safe printable string."""

        if value is None:
            return ""

        return str(value)

    @staticmethod
    def _employee_name(record):
        """Return the employee's full name."""

        employee = record.employee

        return (
            f"{employee.first_name} "
            f"{employee.last_name}"
        ).strip()

    @staticmethod
    def _generated_by_name(user):
        """Return the best available display name for a user."""

        if user is None:
            return "Unknown User"

        first_name = getattr(user, "first_name", None)
        last_name = getattr(user, "last_name", None)

        full_name = " ".join(
            part
            for part in [first_name, last_name]
            if part
        ).strip()

        if full_name:
            return full_name

        return getattr(
            user,
            "username",
            "Unknown User",
        )

    @classmethod
    def generate_payroll_summary_pdf(
        cls,
        selected_period,
        records,
        totals,
        generated_by,
        search_term="",
        selected_department=None,
    ):
        """Generate a landscape PDF payroll-summary report."""

        buffer = BytesIO()

        generated_at = datetime.now()

        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title=(
                f"Payroll Summary - "
                f"{selected_period.period_name}"
            ),
            author=cls.SYSTEM_NAME,
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=22,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#111827"),
            spaceAfter=3,
        )

        company_style = ParagraphStyle(
            "CompanyName",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#1F2937"),
        )

        metadata_style = ParagraphStyle(
            "Metadata",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#4B5563"),
        )

        right_metadata_style = ParagraphStyle(
            "RightMetadata",
            parent=metadata_style,
            alignment=TA_RIGHT,
        )

        summary_label_style = ParagraphStyle(
            "SummaryLabel",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#6B7280"),
        )

        summary_value_style = ParagraphStyle(
            "SummaryValue",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#111827"),
        )

        elements = []

        left_header = [
            Paragraph(
                cls.SYSTEM_NAME,
                company_style,
            ),
            Paragraph(
                "Payroll Summary Report",
                title_style,
            ),
            Paragraph(
                (
                    f"{cls.COMPANY_NAME}<br/>"
                    f"{cls.COMPANY_LOCATION}"
                ),
                metadata_style,
            ),
        ]

        right_header_lines = [
            (
                "<b>Payroll Period:</b> "
                f"{selected_period.period_name}"
            ),
            (
                "<b>Payment Date:</b> "
                f"{selected_period.payment_date.strftime('%d %B %Y')}"
            ),
            (
                "<b>Generated:</b> "
                f"{generated_at.strftime('%d %B %Y %H:%M')}"
            ),
            (
                "<b>Generated By:</b> "
                f"{cls._generated_by_name(generated_by)}"
            ),
        ]

        if selected_department is not None:
            right_header_lines.append(
                (
                    "<b>Department:</b> "
                    f"{selected_department.name}"
                )
            )

        if search_term:
            right_header_lines.append(
                (
                    "<b>Search Filter:</b> "
                    f"{search_term}"
                )
            )

        right_header = Paragraph(
            "<br/>".join(right_header_lines),
            right_metadata_style,
        )

        header_table = Table(
            [[left_header, right_header]],
            colWidths=[
                170 * mm,
                95 * mm,
            ],
        )

        header_table.setStyle(
            TableStyle(
                [
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP",
                    ),
                    (
                        "ALIGN",
                        (1, 0),
                        (1, 0),
                        "RIGHT",
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                ]
            )
        )

        elements.append(header_table)
        elements.append(Spacer(1, 5 * mm))

        summary_data = [
            [
                Paragraph(
                    "Employees",
                    summary_label_style,
                ),
                Paragraph(
                    "Gross Payroll",
                    summary_label_style,
                ),
                Paragraph(
                    "Total Deductions",
                    summary_label_style,
                ),
                Paragraph(
                    "Net Payroll",
                    summary_label_style,
                ),
                Paragraph(
                    "Employer Cost",
                    summary_label_style,
                ),
            ],
            [
                Paragraph(
                    str(totals["employee_count"]),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["gross_pay"]
                    ),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["total_deductions"]
                    ),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["net_pay"]
                    ),
                    summary_value_style,
                ),
                Paragraph(
                    cls._money(
                        totals["employer_cost"]
                    ),
                    summary_value_style,
                ),
            ],
        ]

        summary_table = Table(
            summary_data,
            colWidths=[
                53 * mm,
                53 * mm,
                53 * mm,
                53 * mm,
                53 * mm,
            ],
        )

        summary_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        colors.HexColor("#F3F4F6"),
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.6,
                        colors.HexColor("#D1D5DB"),
                    ),
                    (
                        "INNERGRID",
                        (0, 0),
                        (-1, -1),
                        0.3,
                        colors.HexColor("#E5E7EB"),
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                ]
            )
        )

        elements.append(summary_table)
        elements.append(Spacer(1, 5 * mm))

        table_data = [
            [
                "Employee",
                "Department",
                "Basic",
                "Overtime",
                "Allowances",
                "Gross",
                "NSSA",
                "PAYE",
                "AIDS Levy",
                "Other",
                "Deductions",
                "Net Pay",
            ]
        ]

        for record in records:
            table_data.append(
                [
                    (
                        f"{cls._employee_name(record)}\n"
                        f"{record.employee.employee_number}"
                    ),
                    cls._safe_text(
                        record.employee.department.name
                    ),
                    cls._money(
                        record.basic_salary
                    ),
                    cls._money(
                        record.overtime_amount
                    ),
                    cls._money(
                        record.allowances_total
                    ),
                    cls._money(
                        record.gross_pay
                    ),
                    cls._money(
                        record.nssa
                    ),
                    cls._money(
                        record.paye
                    ),
                    cls._money(
                        record.aids_levy
                    ),
                    cls._money(
                        record.other_deductions_total
                    ),
                    cls._money(
                        record.total_deductions
                    ),
                    cls._money(
                        record.net_pay
                    ),
                ]
            )

        table_data.append(
            [
                "REPORT TOTALS",
                "",
                cls._money(
                    totals["basic_salary"]
                ),
                cls._money(
                    totals["overtime_amount"]
                ),
                cls._money(
                    totals["allowances_total"]
                ),
                cls._money(
                    totals["gross_pay"]
                ),
                cls._money(
                    totals["employee_nssa"]
                ),
                cls._money(
                    totals["paye"]
                ),
                cls._money(
                    totals["aids_levy"]
                ),
                cls._money(
                    totals["other_deductions"]
                ),
                cls._money(
                    totals["total_deductions"]
                ),
                cls._money(
                    totals["net_pay"]
                ),
            ]
        )

        payroll_table = Table(
            table_data,
            repeatRows=1,
            colWidths=[
                35 * mm,
                31 * mm,
                21 * mm,
                19 * mm,
                21 * mm,
                21 * mm,
                18 * mm,
                17 * mm,
                18 * mm,
                17 * mm,
                22 * mm,
                22 * mm,
            ],
        )

        last_row = len(table_data) - 1

        payroll_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1F2937"),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white,
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, 0),
                        7,
                    ),
                    (
                        "ALIGN",
                        (2, 0),
                        (-1, -1),
                        "RIGHT",
                    ),
                    (
                        "ALIGN",
                        (0, 0),
                        (1, -1),
                        "LEFT",
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "FONTNAME",
                        (0, 1),
                        (-1, last_row - 1),
                        "Helvetica",
                    ),
                    (
                        "FONTSIZE",
                        (0, 1),
                        (-1, -1),
                        6.5,
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.35,
                        colors.HexColor("#D1D5DB"),
                    ),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, last_row - 1),
                        [
                            colors.white,
                            colors.HexColor("#F9FAFB"),
                        ],
                    ),
                    (
                        "BACKGROUND",
                        (0, last_row),
                        (-1, last_row),
                        colors.HexColor("#E5E7EB"),
                    ),
                    (
                        "FONTNAME",
                        (0, last_row),
                        (-1, last_row),
                        "Helvetica-Bold",
                    ),
                    (
                        "SPAN",
                        (0, last_row),
                        (1, last_row),
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        3,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        3,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                ]
            )
        )

        elements.append(payroll_table)
        elements.append(Spacer(1, 5 * mm))

        statutory_data = [
            [
                "Employee NSSA",
                cls._money(
                    totals["employee_nssa"]
                ),
                "Employer NSSA",
                cls._money(
                    totals["employer_nssa"]
                ),
                "Total NSSA",
                cls._money(
                    totals["total_nssa"]
                ),
            ],
            [
                "PAYE",
                cls._money(
                    totals["paye"]
                ),
                "AIDS Levy",
                cls._money(
                    totals["aids_levy"]
                ),
                "Total Tax",
                cls._money(
                    totals["total_tax"]
                ),
            ],
        ]

        statutory_table = Table(
            statutory_data,
            colWidths=[
                31 * mm,
                26 * mm,
                31 * mm,
                26 * mm,
                31 * mm,
                26 * mm,
            ],
            hAlign="LEFT",
        )

        statutory_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        colors.HexColor("#F9FAFB"),
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.35,
                        colors.HexColor("#D1D5DB"),
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "Helvetica",
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (0, -1),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTNAME",
                        (2, 0),
                        (2, -1),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTNAME",
                        (4, 0),
                        (4, -1),
                        "Helvetica-Bold",
                    ),
                    (
                        "ALIGN",
                        (1, 0),
                        (1, -1),
                        "RIGHT",
                    ),
                    (
                        "ALIGN",
                        (3, 0),
                        (3, -1),
                        "RIGHT",
                    ),
                    (
                        "ALIGN",
                        (5, 0),
                        (5, -1),
                        "RIGHT",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, -1),
                        7,
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        5,
                    ),
                ]
            )
        )

        elements.append(statutory_table)

        def draw_page_number(canvas, doc):
            """Draw a footer and page number."""

            canvas.saveState()

            page_width, _ = landscape(A4)

            canvas.setStrokeColor(
                colors.HexColor("#D1D5DB")
            )
            canvas.setLineWidth(0.4)

            canvas.line(
                10 * mm,
                8 * mm,
                page_width - 10 * mm,
                8 * mm,
            )

            canvas.setFont(
                "Helvetica",
                7,
            )
            canvas.setFillColor(
                colors.HexColor("#6B7280")
            )

            canvas.drawString(
                10 * mm,
                4.5 * mm,
                (
                    f"{cls.SYSTEM_NAME} | "
                    f"{selected_period.period_name}"
                ),
            )

            canvas.drawRightString(
                page_width - 10 * mm,
                4.5 * mm,
                f"Page {doc.page}",
            )

            canvas.restoreState()

        document.build(
            elements,
            onFirstPage=draw_page_number,
            onLaterPages=draw_page_number,
        )

        buffer.seek(0)

        return buffer

    @classmethod
    def generate_payroll_summary_excel(
        cls,
        selected_period,
        records,
        totals,
        generated_by,
        search_term="",
        selected_department=None,
    ):
        """Generate a formatted Excel payroll-summary workbook."""

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Payroll Summary"

        generated_at = datetime.now()

        dark_fill = PatternFill(
            fill_type="solid",
            fgColor="1F2937",
        )

        blue_fill = PatternFill(
            fill_type="solid",
            fgColor="EAF2FF",
        )

        light_fill = PatternFill(
            fill_type="solid",
            fgColor="F3F4F6",
        )

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="DCFCE7",
        )

        white_font = Font(
            color="FFFFFF",
            bold=True,
        )

        title_font = Font(
            size=18,
            bold=True,
            color="111827",
        )

        subtitle_font = Font(
            size=11,
            bold=True,
            color="374151",
        )

        label_font = Font(
            bold=True,
            color="374151",
        )

        value_font = Font(
            bold=True,
            color="111827",
        )

        thin_side = Side(
            style="thin",
            color="D1D5DB",
        )

        table_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        worksheet.merge_cells(
            "A1:L1"
        )
        worksheet["A1"] = cls.SYSTEM_NAME
        worksheet["A1"].font = title_font
        worksheet["A1"].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        worksheet.merge_cells(
            "A2:L2"
        )
        worksheet["A2"] = "Payroll Summary Report"
        worksheet["A2"].font = Font(
            size=14,
            bold=True,
            color="111827",
        )

        worksheet.merge_cells(
            "A3:L3"
        )
        worksheet["A3"] = (
            f"{cls.COMPANY_NAME} | "
            f"{cls.COMPANY_LOCATION}"
        )
        worksheet["A3"].font = Font(
            italic=True,
            color="6B7280",
        )

        metadata_rows = [
            (
                "Payroll Period",
                selected_period.period_name,
            ),
            (
                "Payment Date",
                selected_period.payment_date.strftime(
                    "%d %B %Y"
                ),
            ),
            (
                "Generated",
                generated_at.strftime(
                    "%d %B %Y %H:%M"
                ),
            ),
            (
                "Generated By",
                cls._generated_by_name(
                    generated_by
                ),
            ),
        ]

        if selected_department is not None:
            metadata_rows.append(
                (
                    "Department",
                    selected_department.name,
                )
            )

        if search_term:
            metadata_rows.append(
                (
                    "Search Filter",
                    search_term,
                )
            )

        metadata_start_row = 5

        for offset, (label, value) in enumerate(
            metadata_rows
        ):
            row = metadata_start_row + offset

            worksheet.cell(
                row=row,
                column=1,
                value=label,
            )
            worksheet.cell(
                row=row,
                column=2,
                value=value,
            )

            worksheet.cell(
                row=row,
                column=1,
            ).font = label_font

        summary_start_row = (
            metadata_start_row
            + len(metadata_rows)
            + 2
        )

        summary_headers = [
            "Employees",
            "Gross Payroll",
            "Total Deductions",
            "Net Payroll",
            "Employer Cost",
        ]

        summary_values = [
            totals["employee_count"],
            totals["gross_pay"],
            totals["total_deductions"],
            totals["net_pay"],
            totals["employer_cost"],
        ]

        for column, heading in enumerate(
            summary_headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=summary_start_row,
                column=column,
                value=heading,
            )

            cell.fill = blue_fill
            cell.font = label_font
            cell.border = table_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            value_cell = worksheet.cell(
                row=summary_start_row + 1,
                column=column,
                value=summary_values[column - 1],
            )

            value_cell.fill = light_fill
            value_cell.font = value_font
            value_cell.border = table_border
            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            if column > 1:
                value_cell.number_format = (
                    '$#,##0.00'
                )

        table_header_row = summary_start_row + 4

        headings = [
            "Employee Number",
            "Employee",
            "Department",
            "Basic Salary",
            "Overtime",
            "Allowances",
            "Gross Pay",
            "NSSA",
            "PAYE",
            "AIDS Levy",
            "Other Deductions",
            "Total Deductions",
            "Net Pay",
            "Employer NSSA",
            "Employer Cost",
        ]

        for column, heading in enumerate(
            headings,
            start=1,
        ):
            cell = worksheet.cell(
                row=table_header_row,
                column=column,
                value=heading,
            )

            cell.fill = dark_fill
            cell.font = white_font
            cell.border = table_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        data_start_row = table_header_row + 1

        for row_offset, record in enumerate(
            records
        ):
            row = data_start_row + row_offset

            values = [
                record.employee.employee_number,
                cls._employee_name(record),
                record.employee.department.name,
                record.basic_salary,
                record.overtime_amount,
                record.allowances_total,
                record.gross_pay,
                record.nssa,
                record.paye,
                record.aids_levy,
                record.other_deductions_total,
                record.total_deductions,
                record.net_pay,
                record.employer_nssa,
                record.employer_cost,
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row,
                    column=column,
                    value=value,
                )

                cell.border = table_border
                cell.alignment = Alignment(
                    vertical="center",
                )

                if column >= 4:
                    cell.number_format = (
                        '$#,##0.00'
                    )
                    cell.alignment = Alignment(
                        horizontal="right",
                        vertical="center",
                    )

        totals_row = (
            data_start_row
            + len(records)
        )

        worksheet.cell(
            row=totals_row,
            column=1,
            value="REPORT TOTALS",
        )

        worksheet.merge_cells(
            start_row=totals_row,
            start_column=1,
            end_row=totals_row,
            end_column=3,
        )

        totals_mapping = {
            4: totals["basic_salary"],
            5: totals["overtime_amount"],
            6: totals["allowances_total"],
            7: totals["gross_pay"],
            8: totals["employee_nssa"],
            9: totals["paye"],
            10: totals["aids_levy"],
            11: totals["other_deductions"],
            12: totals["total_deductions"],
            13: totals["net_pay"],
            14: totals["employer_nssa"],
            15: totals["employer_cost"],
        }

        for column in range(
            1,
            len(headings) + 1,
        ):
            cell = worksheet.cell(
                row=totals_row,
                column=column,
            )

            cell.fill = green_fill
            cell.font = Font(
                bold=True,
                color="111827",
            )
            cell.border = table_border
            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column == 1
                    else "right"
                ),
                vertical="center",
            )

            if column in totals_mapping:
                cell.value = totals_mapping[column]
                cell.number_format = '$#,##0.00'

        statutory_start_row = totals_row + 3

        statutory_rows = [
            (
                "Employee NSSA",
                totals["employee_nssa"],
            ),
            (
                "Employer NSSA",
                totals["employer_nssa"],
            ),
            (
                "Total NSSA",
                totals["total_nssa"],
            ),
            (
                "PAYE",
                totals["paye"],
            ),
            (
                "AIDS Levy",
                totals["aids_levy"],
            ),
            (
                "Total Tax",
                totals["total_tax"],
            ),
        ]

        worksheet.cell(
            row=statutory_start_row,
            column=1,
            value="Statutory Summary",
        )

        worksheet.cell(
            row=statutory_start_row,
            column=1,
        ).font = subtitle_font

        for offset, (label, value) in enumerate(
            statutory_rows,
            start=1,
        ):
            row = statutory_start_row + offset

            label_cell = worksheet.cell(
                row=row,
                column=1,
                value=label,
            )

            value_cell = worksheet.cell(
                row=row,
                column=2,
                value=value,
            )

            label_cell.border = table_border
            label_cell.font = label_font
            label_cell.fill = light_fill

            value_cell.border = table_border
            value_cell.number_format = '$#,##0.00'
            value_cell.alignment = Alignment(
                horizontal="right",
            )

        column_widths = {
            1: 18,
            2: 25,
            3: 25,
            4: 16,
            5: 14,
            6: 15,
            7: 15,
            8: 13,
            9: 13,
            10: 14,
            11: 18,
            12: 18,
            13: 15,
            14: 16,
            15: 17,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = width

        worksheet.freeze_panes = (
            f"A{data_start_row}"
        )

        worksheet.auto_filter.ref = (
            f"A{table_header_row}:"
            f"O{totals_row - 1}"
        )

        worksheet.sheet_view.showGridLines = False

        worksheet.page_setup.orientation = (
            "landscape"
        )
        worksheet.page_setup.paperSize = (
            worksheet.PAPERSIZE_A4
        )
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        worksheet.print_title_rows = (
            f"{table_header_row}:{table_header_row}"
        )

        worksheet.freeze_panes = (
            f"A{data_start_row}"
        )

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return buffer
